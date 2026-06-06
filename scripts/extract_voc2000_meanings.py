from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def extract_rows(xlsx_path: Path) -> list[dict[str, str]]:
    ws = load_workbook(xlsx_path, data_only=True)["工作表1"]

    merged_starts = {
        rng.min_row: rng.max_row
        for rng in ws.merged_cells.ranges
        if rng.min_col == 1 and rng.max_col == 1
    }

    rows: list[dict[str, str]] = []
    row = 2
    while row <= ws.max_row:
        end_row = merged_starts.get(row, row)
        word = ws[f"A{row}"].value
        if word:
            meaning_parts = [
                str(ws[f"C{current_row}"].value).strip()
                for current_row in range(row, end_row + 1)
                if ws[f"C{current_row}"].value not in (None, "")
            ]
            rows.append(
                {
                    "word": str(word).strip(),
                    "meaning": "".join(meaning_parts).strip(),
                }
            )
        row = end_row + 1

    return rows


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("voc2000.xlsx")
    output_path = (
        Path(sys.argv[2]) if len(sys.argv) > 2 else Path("outputs") / "voc2000_meanings.json"
    )

    rows = extract_rows(xlsx_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Rows: {len(rows)}")
    print(f"Saved {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
