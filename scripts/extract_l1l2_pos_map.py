from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


REPO_ROOT = Path(__file__).resolve().parent.parent


def normalize(word: str) -> str:
    return re.sub(r"\s+", " ", word.strip().lower().replace("／", "/").replace("’", "'"))


def extract_from_voc2000() -> dict[str, str]:
    ws = load_workbook(REPO_ROOT / "voc2000.xlsx", data_only=True)["工作表1"]
    merged_starts = {
        rng.min_row: rng.max_row
        for rng in ws.merged_cells.ranges
        if rng.min_col == 1 and rng.max_col == 1
    }
    result: dict[str, str] = {}
    row = 2
    while row <= ws.max_row:
        end_row = merged_starts.get(row, row)
        word = ws[f"A{row}"].value
        pos = ws[f"B{row}"].value
        if word and pos:
            result[normalize(str(word))] = str(pos).strip()
        row = end_row + 1
    return result


PDF_RE = re.compile(
    r"^\s*\d+\s+(.+?)\s+(名詞|動詞|形容詞|副詞|介系詞|代名詞|限定詞|連接詞|助動詞|感嘆詞)\s+(.+?)\s*$"
)


def extract_from_juniorhigh_pdf() -> dict[str, str]:
    reader = PdfReader(str(REPO_ROOT / "國中單字王 2000 .pdf"))
    result: dict[str, str] = {}
    for page in reader.pages:
        for line in (page.extract_text() or "").splitlines():
            match = PDF_RE.match(line.strip())
            if match:
                result[normalize(match.group(1))] = match.group(2)
    return result


def main() -> int:
    combined = extract_from_voc2000()
    combined.update(extract_from_juniorhigh_pdf())
    output_path = REPO_ROOT / "outputs" / "l1l2_pos_map.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(combined, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Rows: {len(combined)}")
    print(f"Saved {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
