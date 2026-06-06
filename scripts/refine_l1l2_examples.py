from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "詞彙分級表L1L2.xlsx"
BACKUP_PATH = ROOT / "詞彙分級表L1L2_before_refine.xlsx"
SOURCE_EXAMPLES_PATH = ROOT / "outputs" / "vocab_examples" / "英文單字7000_意思例句.xlsx"


BAD_MARKERS = (
    "useful English word",
    "seem yummy",
    "seem yucky",
    "drank some sailor",
    "had bakery",
    "looks able",
    "looks absent",
    "We will talk about it in today",
    " in today.",
)

GENERIC_MARKERS = (
    "The teacher gave us an example with",
    "This is a simple sentence with",
    "Please use the word",
    "We learned about",
    "We discussed",
    "was important to the story",
    "appeared in the story",
)


EXACT_EXAMPLES = {
    "yucky": "The spoiled milk smelled yucky. (那瓶壞掉的牛奶聞起來很噁心。)",
    "yummy": "The cookies tasted yummy. (那些餅乾吃起來很好吃。)",
    "today": "Today is my birthday. (今天是我的生日。)",
    "tomorrow": "We will visit the museum tomorrow. (我們明天會參觀博物館。)",
    "yesterday": "I finished the book yesterday. (我昨天讀完了那本書。)",
    "baby": "The baby smiled at her mother. (那個嬰兒對媽媽微笑。)",
    "beginner": "The beginner learned the first lesson quickly. (那位初學者很快學會第一課。)",
    "blackboard": "The teacher wrote the answer on the blackboard. (老師把答案寫在黑板上。)",
    "bloody": "He cleaned the bloody towel carefully. (他小心地清洗那條沾血的毛巾。)",
    "bony": "The stray dog looked thin and bony. (那隻流浪狗看起來又瘦又骨瘦如柴。)",
    "christmas": "We exchange gifts at Christmas. (我們在聖誕節交換禮物。)",
    "christmas/xmas": "We exchange gifts at Christmas. (我們在聖誕節交換禮物。)",
    "good-bye": "She waved and said good-bye at the door. (她在門口揮手說再見。)",
    "goodbye": "She waved and said goodbye at the door. (她在門口揮手說再見。)",
    "good-by": "She waved and said good-by at the door. (她在門口揮手說再見。)",
    "bye-bye": "The child said bye-bye to his father. (那個孩子向爸爸說再見。)",
    "bye": "The child said bye to his father. (那個孩子向爸爸說再見。)",
    "ahead": "Look ahead before you cross the road. (過馬路前要向前看。)",
    "almost": "I almost missed the bus this morning. (我今天早上差點錯過公車。)",
    "already": "She has already finished her homework. (她已經完成作業了。)",
    "also": "My brother also likes this song. (我哥哥也喜歡這首歌。)",
    "always": "He always brushes his teeth before bed. (他睡前總是刷牙。)",
    "among": "The teacher stood among the students. (老師站在學生們之間。)",
    "another": "May I have another cup of water? (我可以再喝一杯水嗎？)",
    "any": "Any student can join the club. (任何學生都可以加入這個社團。)",
    "anything": "You can ask me anything about the lesson. (你可以問我任何關於這課的問題。)",
    "anybody": "Anybody can learn with enough practice. (任何人只要充分練習都能學會。)",
    "anyone": "Anyone can join the game after school. (任何人放學後都可以加入遊戲。)",
    "altogether": "There are thirty students altogether. (總共有三十位學生。)",
    "anyhow": "Anyhow, we still finished the work. (無論如何，我們還是完成了工作。)",
    "anyway": "Anyway, thank you for your help. (無論如何，謝謝你的幫忙。)",
    "around": "We walked around the lake after lunch. (我們午餐後繞著湖散步。)",
    "away": "Please put your phone away during class. (上課時請把手機收起來。)",
    "backward": "The little boy walked backward for fun. (那個小男孩為了好玩倒著走。)",
    "backwards": "The little boy walked backwards for fun. (那個小男孩為了好玩倒著走。)",
    "air": "Fresh air came in through the window. (新鮮空氣從窗戶吹進來。)",
    "apple": "She ate an apple after lunch. (她午餐後吃了一顆蘋果。)",
    "anger": "He tried to hide his anger. (他試著隱藏自己的怒氣。)",
    "animal": "The animal ran back into the forest. (那隻動物跑回森林裡。)",
    "army": "His brother joined the army last year. (他的哥哥去年從軍。)",
    "art": "The museum is full of modern art. (那間博物館充滿現代藝術作品。)",
    "bank": "She put her savings in the bank. (她把存款放在銀行裡。)",
    "cocoa": "I like to drink hot cocoa in winter. (我喜歡在冬天喝熱可可。)",
    "common": "This is a common mistake for beginners. (這是初學者常犯的錯誤。)",
    "do": "Please do your homework before dinner. (請在晚餐前做完作業。)",
    "english": "We practice English every morning. (我們每天早上練習英文。)",
    "ghost": "The children told a story about a ghost. (孩子們講了一個關於鬼魂的故事。)",
    "god": "The temple honors an ancient god. (這座廟供奉一位古老的神。)",
    "goddess": "The story is about a brave goddess. (這個故事是關於一位勇敢的女神。)",
    "glasses": "I need my glasses to read the book. (我需要眼鏡才能讀這本書。)",
    "chopstick": "She picked up the noodles with chopsticks. (她用筷子夾起麵條。)",
    "chopsticks": "She picked up the noodles with chopsticks. (她用筷子夾起麵條。)",
    "confucius": "Confucius taught people to value learning. (孔子教導人們重視學習。)",
    "congratulation": "She sent a card of congratulation to her friend. (她寄了一張祝賀卡給朋友。)",
    "congratulations": "Congratulations on winning the race! (恭喜你贏得比賽！)",
    "contain": "This box contains old family photos. (這個盒子裡裝著舊的家庭照片。)",
    "confirm": "Please confirm the meeting time with me. (請和我確認會議時間。)",
    "cough": "Cover your mouth when you cough. (咳嗽時請遮住嘴巴。)",
    "costly": "The repair was costly, but it was necessary. (修理費很昂貴，但那是必要的。)",
    "ask": "Please ask your teacher for help. (請向老師求助。)",
    "bathe": "The baby likes to bathe before bed. (那個嬰兒喜歡睡前洗澡。)",
    "begin": "The show will begin at seven. (表演將在七點開始。)",
    "buy": "I want to buy a new notebook. (我想買一本新的筆記本。)",
    "care": "We care about our friends. (我們關心我們的朋友。)",
    "carry": "Can you carry this box for me? (你可以幫我搬這個箱子嗎？)",
    "close": "Please close the door quietly. (請安靜地關門。)",
    "cut": "She cut the paper with scissors. (她用剪刀剪紙。)",
    "draw": "He likes to draw animals. (他喜歡畫動物。)",
    "drive": "My father can drive a bus. (我爸爸會開公車。)",
    "hear": "I can hear music from the classroom. (我可以從教室聽到音樂。)",
    "join": "Would you like to join our team? (你想加入我們的隊伍嗎？)",
    "lay": "Please lay the book on the desk. (請把書放在桌上。)",
    "leave": "We will leave after lunch. (我們午餐後會離開。)",
    "lie": "Do not lie to your parents. (不要對父母說謊。)",
    "listen": "Please listen to the teacher carefully. (請仔細聽老師說話。)",
    "look": "Look at the picture on the wall. (看牆上的那張圖。)",
    "open": "Please open the window. (請打開窗戶。)",
    "put": "Put your bag under the desk. (把你的書包放在桌子下面。)",
    "read": "She likes to read stories before bed. (她喜歡睡前讀故事。)",
    "see": "I can see the mountains from here. (我從這裡可以看見山。)",
    "shut": "Please shut the gate after you leave. (你離開後請關上大門。)",
    "speak": "He can speak English very well. (他英文說得很好。)",
    "spell": "Can you spell your name for me? (你可以把你的名字拼給我聽嗎？)",
    "start": "The race will start in five minutes. (比賽五分鐘後開始。)",
    "take": "Please take an umbrella with you. (請帶一把傘出門。)",
    "visit": "We will visit our grandparents this weekend. (我們這個週末會拜訪祖父母。)",
    "wash": "Please wash your hands before dinner. (晚餐前請洗手。)",
    "watch": "We watch the news after dinner. (我們晚餐後看新聞。)",
    "write": "Please write your name on the paper. (請把你的名字寫在紙上。)",
    "boil": "Boil the water before you make tea. (泡茶前先把水煮沸。)",
    "divide": "The teacher will divide the class into groups. (老師會把全班分成小組。)",
    "greet": "The students greet their teacher every morning. (學生們每天早上向老師問好。)",
    "increase": "Exercise can increase your strength. (運動可以增強你的力氣。)",
    "danger": "The sign warned us of danger. (那個標誌警告我們有危險。)",
    "difficulty": "She had difficulty solving the problem. (她解那道題目時遇到困難。)",
    "importance": "The teacher explained the importance of practice. (老師說明練習的重要性。)",
    "possibility": "There is a possibility of rain tonight. (今晚有下雨的可能性。)",
    "believable": "Her story sounded believable. (她的故事聽起來可信。)",
    "elder": "My elder sister studies in Taipei. (我的姐姐在台北讀書。)",
    "fearful": "The child felt fearful in the dark room. (那個孩子在黑暗的房間裡感到害怕。)",
    "following": "Please read the following sentence. (請閱讀接下來的句子。)",
    "grassy": "The children played on a grassy hill. (孩子們在長滿草的小山上玩。)",
    "hateful": "He regretted saying those hateful words. (他後悔說了那些可恨的話。)",
    "lone": "A lone tree stood beside the road. (一棵孤零零的樹立在路旁。)",
    "lower": "Please move the picture to a lower place. (請把那張圖移到較低的位置。)",
    "measurable": "The plan has measurable results. (這個計畫有可衡量的成果。)",
    "military": "The military base is near the coast. (那座軍事基地靠近海岸。)",
    "movable": "The table is light and movable. (這張桌子很輕，也可以移動。)",
    "naked": "The baby was naked after the bath. (那個嬰兒洗完澡後光著身子。)",
    "official": "The official report came out yesterday. (官方報告昨天公布了。)",
    "plain": "She wore a plain white shirt. (她穿了一件樸素的白襯衫。)",
    "rapid": "The city has seen rapid change. (這座城市經歷了快速的變化。)",
    "rocky": "The road became rocky near the mountain. (靠近山區時道路變得多岩石。)",
    "salty": "This soup is too salty for me. (這碗湯對我來說太鹹了。)",
    "southern": "They live in the southern part of Taiwan. (他們住在台灣南部。)",
    "moment": "Wait a moment, please. (請等一下。)",
    "most": "Most students finished the homework. (大多數學生完成了作業。)",
    "sorry": "I am sorry for being late. (我很抱歉遲到了。)",
    "he": "He plays basketball after school. (他放學後打籃球。)",
    "him": "I gave him a book. (我給了他一本書。)",
    "his": "His backpack is under the desk. (他的背包在桌子下面。)",
    "she": "She reads a story every night. (她每天晚上讀一個故事。)",
    "her": "I saw her at the library. (我在圖書館看見她。)",
    "hers": "This pencil is hers. (這枝鉛筆是她的。)",
    "it": "It is raining outside. (外面正在下雨。)",
    "its": "The dog wagged its tail. (那隻狗搖著牠的尾巴。)",
    "i": "I like English class. (我喜歡英文課。)",
    "me": "Please wait for me. (請等我。)",
    "my": "My notebook is on the desk. (我的筆記本在桌上。)",
    "mine": "This seat is mine. (這個座位是我的。)",
    "we": "We study together after school. (我們放學後一起讀書。)",
    "us": "The teacher gave us a new book. (老師給了我們一本新書。)",
    "our": "Our classroom is clean. (我們的教室很乾淨。)",
    "ours": "This garden is ours. (這座花園是我們的。)",
    "you": "You can sit here. (你可以坐在這裡。)",
    "your": "Your answer is correct. (你的答案是正確的。)",
    "yours": "Is this umbrella yours? (這把傘是你的嗎？)",
    "they": "They play soccer on Sundays. (他們星期天踢足球。)",
    "them": "I met them at the station. (我在車站遇見他們。)",
    "their": "Their house is near the park. (他們的房子在公園附近。)",
    "theirs": "The red bicycle is theirs. (那輛紅色腳踏車是他們的。)",
    "who": "Who is knocking on the door? (誰在敲門？)",
    "whom": "Whom did you invite to dinner? (你邀請了誰來吃晚餐？)",
    "whose": "Whose jacket is this? (這是誰的外套？)",
    "day": "It was a sunny day. (那是晴朗的一天。)",
    "grandfather": "My grandfather tells funny stories. (我的祖父會講有趣的故事。)",
    "grandpa": "My grandpa tells funny stories. (我的爺爺會講有趣的故事。)",
    "grandmother": "My grandmother grows flowers in the garden. (我的祖母在花園裡種花。)",
    "grandma": "My grandma grows flowers in the garden. (我的奶奶在花園裡種花。)",
    "grandchild": "The old woman smiled at her grandchild. (那位老太太對她的孫子微笑。)",
    "granddaughter": "His granddaughter drew a picture for him. (他的孫女畫了一張圖給他。)",
    "grandson": "Her grandson visits her every weekend. (她的孫子每個週末都去看她。)",
    "grass": "The children sat on the grass. (孩子們坐在草地上。)",
    "gray": "The sky turned gray before the rain. (下雨前天空變成灰色。)",
    "grey": "The sky turned grey before the rain. (下雨前天空變成灰色。)",
    "haircut": "He got a short haircut before school started. (開學前他剪了一個短髮。)",
    "hello": "She said hello to her new classmate. (她向新同學說哈囉。)",
    "hill": "We climbed a small hill after lunch. (我們午餐後爬上一座小山丘。)",
    "history": "History helps us understand the past. (歷史幫助我們了解過去。)",
    "hole": "There is a small hole in my sock. (我的襪子上有一個小洞。)",
    "holiday": "We visited our cousins during the holiday. (假日期間我們拜訪了表兄弟姊妹。)",
    "homework": "I finished my homework before dinner. (我在晚餐前完成了功課。)",
    "hour": "The meeting lasted one hour. (會議持續了一個小時。)",
    "house": "Their house is near the river. (他們的房子在河邊。)",
    "how": "How did you solve the problem? (你是怎麼解決這個問題的？)",
    "in": "The keys are in my bag. (鑰匙在我的包包裡。)",
    "inch": "The ruler is twelve inches long. (這把尺有十二英寸長。)",
    "interest": "She has a strong interest in music. (她對音樂有濃厚的興趣。)",
    "is": "The soup is hot. (這碗湯是熱的。)",
    "jam": "I spread strawberry jam on the toast. (我把草莓果醬抹在吐司上。)",
    "chess": "My uncle taught me how to play chess. (我叔叔教我下西洋棋。)",
    "couple": "The couple walked along the beach. (那對夫妻沿著海灘散步。)",
    "court": "The players practiced on the basketball court. (球員們在籃球場練習。)",
    "crab": "We saw a crab on the beach. (我們在海灘上看到一隻螃蟹。)",
    "crane": "A crane stood quietly by the pond. (一隻鶴安靜地站在池塘邊。)",
    "crayon": "The child drew a house with a red crayon. (那個孩子用紅色蠟筆畫了一間房子。)",
    "cream": "She put cream on the cake. (她把鮮奶油放在蛋糕上。)",
    "crime": "The police are investigating the crime. (警方正在調查那起犯罪事件。)",
    "crisis": "The team stayed calm during the crisis. (團隊在危機中保持冷靜。)",
    "crop": "The farmer checked the crop after the storm. (暴風雨後農夫檢查作物。)",
    "crow": "The rooster began to crow at dawn. (公雞在黎明時開始啼叫。)",
    "culture": "Food is an important part of culture. (食物是文化的重要部分。)",
    "custom": "Giving red envelopes is a New Year custom. (包紅包是新年的習俗。)",
    "customer": "The customer asked for a receipt. (那位顧客要求一張收據。)",
    "damage": "The typhoon caused serious damage. (颱風造成嚴重損害。)",
    "data": "The scientist checked the data twice. (那位科學家檢查了兩次數據。)",
    "dawn": "We left home at dawn. (我們在黎明時離開家。)",
    "debate": "The students had a debate about school rules. (學生們針對校規進行辯論。)",
    "debt": "He worked hard to pay off his debt. (他努力工作來還清債務。)",
    "decision": "She made an important decision yesterday. (她昨天做了一個重要決定。)",
    "ice": "The ice melted in the sun. (冰在陽光下融化了。)",
    "joke": "He told a funny joke at lunch. (他午餐時講了一個好笑的笑話。)",
    "joy": "The children shouted with joy. (孩子們高興地大叫。)",
    "just": "I just finished my homework. (我剛剛完成作業。)",
    "lamp": "The lamp beside my bed is bright. (我床邊的燈很亮。)",
    "leaf": "A yellow leaf fell from the tree. (一片黃葉從樹上落下。)",
    "lily": "She put a white lily in the vase. (她把一朵白百合插進花瓶。)",
    "lip": "He cut his lip during the game. (他在比賽中割傷了嘴唇。)",
    "lot": "A lot of students joined the trip. (許多學生參加了這趟旅行。)",
    "machine": "The machine makes bread every morning. (那台機器每天早上製作麵包。)",
    "mail": "I checked the mail after school. (我放學後查看郵件。)",
    "man": "The man helped us find the station. (那位男士幫我們找到車站。)",
    "many": "Many people came to the concert. (許多人來聽演唱會。)",
    "map": "We used a map to find the museum. (我們用地圖找到博物館。)",
    "maybe": "Maybe we can go hiking tomorrow. (也許我們明天可以去健行。)",
    "mile": "We walked one mile after dinner. (我們晚餐後走了一英里。)",
    "money": "She saved money for a new bike. (她存錢買新腳踏車。)",
    "morning": "I drink milk every morning. (我每天早上喝牛奶。)",
    "mother": "My mother cooks dinner for us. (我媽媽為我們煮晚餐。)",
    "mountain": "The mountain is covered with snow. (那座山覆蓋著雪。)",
    "degree": "The temperature rose by three degrees. (溫度上升了三度。)",
    "delay": "The heavy rain caused a delay. (大雨造成了延誤。)",
    "dentist": "I went to the dentist for a checkup. (我去牙醫那裡做檢查。)",
    "deny": "He did not deny his mistake. (他沒有否認自己的錯誤。)",
    "depth": "The lake has great depth. (那座湖很深。)",
    "desert": "A camel can live in the desert. (駱駝可以生活在沙漠裡。)",
    "desire": "She has a strong desire to learn. (她有強烈的學習慾望。)",
    "dessert": "We had fruit for dessert. (我們飯後甜點吃水果。)",
    "diamond": "The ring has a small diamond. (那枚戒指上有一顆小鑽石。)",
    "diary": "She writes in her diary every night. (她每天晚上寫日記。)",
    "dictionary": "I looked up the word in a dictionary. (我在字典裡查了那個字。)",
    "difference": "Can you see the difference between them? (你看得出它們之間的差異嗎？)",
    "dinosaur": "The museum has a dinosaur model. (那間博物館有一個恐龍模型。)",
    "direction": "Please follow the direction on the sign. (請照著標誌上的方向走。)",
    "discussion": "We had a short discussion after class. (我們下課後進行了簡短討論。)",
    "display": "The store put new shoes on display. (那家店展示了新鞋。)",
    "distance": "The distance between the towns is short. (兩個城鎮之間的距離很短。)",
    "donkey": "The donkey carried bags up the hill. (那頭驢把袋子馱上山。)",
    "dot": "She drew a red dot on the paper. (她在紙上畫了一個紅點。)",
    "doughnut": "He bought a doughnut at the bakery. (他在麵包店買了一個甜甜圈。)",
    "mr.": "Mr. Chen teaches math at our school. (陳先生在我們學校教數學。)",
    "mister": "Mister Chen teaches math at our school. (陳先生在我們學校教數學。)",
    "mrs.": "Mrs. Lin lives next door. (林太太住在隔壁。)",
    "ms.": "Ms. Wang is our new teacher. (王女士是我們的新老師。)",
    "sir": "Sir, may I ask a question? (先生，我可以問一個問題嗎？)",
    "no": "No, I do not need any help. (不，我不需要任何幫忙。)",
    "nope": "Nope, I have not seen your keys. (沒有，我沒有看到你的鑰匙。)",
    "o.k.": "Is it O.K. to sit here? (坐在這裡可以嗎？)",
    "ok": "Is it OK to sit here? (坐在這裡可以嗎？)",
    "okay": "It is okay to make mistakes. (犯錯是可以的。)",
    "so": "I was tired, so I went to bed early. (我很累，所以早早上床睡覺。)",
    "this": "This book belongs to me. (這本書是我的。)",
    "those": "Those shoes are too small. (那些鞋子太小了。)",
    "thus": "He studied hard and thus passed the test. (他努力讀書，因此通過了考試。)",
    "together": "We walked to school together. (我們一起走路去學校。)",
    "tonight": "I will call you tonight. (我今晚會打電話給你。)",
    "too": "This box is too heavy for me. (這個箱子對我來說太重了。)",
    "twice": "I read the story twice. (我把故事讀了兩次。)",
    "zero": "The score was zero to zero. (比數是零比零。)",
    "per": "The car can travel sixty miles per hour. (這輛車每小時可行駛六十英里。)",
    "nearly": "We nearly missed the train. (我們差點錯過火車。)",
    "hardly": "She hardly slept last night. (她昨晚幾乎沒睡。)",
    "simply": "He explained the rule simply. (他簡單地解釋了規則。)",
    "govern": "Good rules help govern a country. (好的規則有助於治理國家。)",
    "hum": "She began to hum a quiet song. (她開始哼一首安靜的歌。)",
    "indicate": "The arrow indicates the right direction. (箭頭指出正確方向。)",
    "maintain": "We should maintain a healthy habit. (我們應該保持健康的習慣。)",
    "influence": "Friends can influence our choices. (朋友會影響我們的選擇。)",
    "interview": "The reporter will interview the singer. (記者將訪問那位歌手。)",
    "pat": "She gave the dog a gentle pat. (她輕輕拍了那隻狗一下。)",
    "pitch": "He pitched the ball across the field. (他把球投過球場。)",
    "pose": "The model posed for a photo. (那位模特兒擺姿勢拍照。)",
    "propose": "They will propose a new plan tomorrow. (他們明天會提出新計畫。)",
    "regard": "Teachers regard safety as important. (老師們重視安全。)",
    "settle": "They decided to settle in the small town. (他們決定在那個小鎮定居。)",
    "sink": "The stone will sink in water. (石頭會沉入水中。)",
    "rush": "Do not rush across the street. (不要急著衝過馬路。)",
    "however": "However, we decided to try again. (然而，我們決定再試一次。)",
    "therefore": "It rained; therefore, the game was canceled. (下雨了，因此比賽取消了。)",
    "used to": "My father used to ride a bike to school. (我爸爸過去常騎腳踏車上學。)",
    "whenever": "Call me whenever you need help. (無論何時你需要幫忙都可以打給我。)",
    "wherever": "She carries a notebook wherever she goes. (她無論去哪裡都帶著筆記本。)",
    "whoever": "Whoever finishes first can choose the game. (無論誰先完成都可以選遊戲。)",
    "widen": "They plan to widen the road next year. (他們計畫明年拓寬那條路。)",
    "wed": "They will wed in a small church. (他們將在一間小教堂結婚。)",
    "flash": "Lightning flashed across the sky. (閃電劃過天空。)",
    "stretch": "Please stretch before you run. (跑步前請先伸展。)",
    "supply": "The school will supply clean water. (學校會供應乾淨的水。)",
    "tear": "Be careful not to tear the paper. (小心不要撕破紙。)",
    "tip": "He left a tip for the waiter. (他留了小費給服務生。)",
    "trade": "The two children trade stickers. (那兩個孩子交換貼紙。)",
    "whisper": "Please whisper in the library. (在圖書館請輕聲說話。)",
    "third": "She won third place in the race. (她在比賽中得到第三名。)",
    "pound": "The bag weighs one pound. (這個袋子重一磅。)",
    "meter": "The table is one meter long. (這張桌子長一公尺。)",
    "teen": "The teen helped his neighbor carry bags. (那位青少年幫鄰居提袋子。)",
    "teens": "Many teens enjoy music and sports. (許多青少年喜歡音樂和運動。)",
    "teenage": "Teenage years can bring many changes. (青少年時期會帶來許多變化。)",
    "teenager": "The teenager joined the school team. (那位青少年加入了校隊。)",
    "t-shirt": "He wore a blue T-shirt to school. (他穿了一件藍色 T 恤去學校。)",
    "noon": "We eat lunch at noon. (我們中午吃午餐。)",
    "pencil": "I wrote the answer with a pencil. (我用鉛筆寫下答案。)",
    "pet": "My pet sleeps beside my bed. (我的寵物睡在我的床邊。)",
    "rainbow": "A rainbow appeared after the rain. (雨後出現了一道彩虹。)",
    "sand": "The children played in the sand. (孩子們在沙中玩耍。)",
    "silver": "She wore a silver ring. (她戴了一枚銀戒指。)",
    "taxicab": "We took a taxicab to the station. (我們搭計程車去車站。)",
    "taxi": "We took a taxi to the station. (我們搭計程車去車站。)",
    "cab": "We took a cab to the station. (我們搭計程車去車站。)",
    "tree": "A tall tree stands near my house. (一棵高樹立在我家附近。)",
    "weekend": "We visited the museum on the weekend. (我們週末參觀了博物館。)",
    "mass": "A mass of people waited outside. (一大群人在外面等候。)",
    "stream": "A clear stream runs through the valley. (一條清澈的小溪流過山谷。)",
    "stress": "Exercise helps me reduce stress. (運動幫助我減輕壓力。)",
    "struggle": "Her struggle made her stronger. (她的奮鬥讓她更堅強。)",
    "thunder": "Thunder woke me up last night. (雷聲昨晚把我吵醒。)",
    "violin": "She plays the violin after school. (她放學後拉小提琴。)",
}


def normalize(text: object) -> str:
    value = str(text or "").strip().lower()
    value = re.sub(r"\s*\(\d+\)", "", value)
    value = re.sub(r"\([^)]*\)", "", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def word_forms(word: object) -> list[str]:
    base = normalize(word)
    forms = [base]
    for part in re.split(r"[/;]", base):
        part = part.strip()
        if part:
            forms.append(part)
    if forms:
        forms.append(forms[0].replace("(s)", "s"))
        forms.append(forms[0].replace("(s)", ""))
    return list(dict.fromkeys(forms))


def surface_word(word: object) -> str:
    form = word_forms(word)[0]
    return re.sub(r"\(s\)", "", form.split("/")[0]).strip()


def first_gloss(meaning: object) -> str:
    text = str(meaning or "").strip()
    text = re.sub(r"\[[^\]]*\]", "", text)
    text = re.sub(r"（[^）]*）", "", text)
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"[A-Za-z].*$", "", text)
    text = re.split(r"[；;，,、/（(]", text)[0].strip()
    return text or "這個詞"


def clean_adj_cn(gloss: str) -> str:
    return gloss[:-1] if gloss.endswith("的") else gloss


def article_for(word: str) -> str:
    return "an" if re.match(r"^[aeiou]", word) else "a"


def has_any(text: str, keywords: tuple[str, ...]) -> bool:
    return any(keyword in text for keyword in keywords)


def is_bad(example: object) -> bool:
    text = str(example or "")
    return any(marker in text for marker in BAD_MARKERS)


def is_generic(example: object) -> bool:
    text = str(example or "")
    return any(marker in text for marker in GENERIC_MARKERS)


def load_source_examples() -> dict[str, str]:
    source = {}
    wb = load_workbook(SOURCE_EXAMPLES_PATH, data_only=True)
    for sheet in wb.worksheets:
        for row in range(2, sheet.max_row + 1):
            word = sheet.cell(row, 1).value
            example = sheet.cell(row, 3).value
            if not word or not example:
                continue
            for form in word_forms(word):
                source.setdefault(form, str(example).strip())
    return source


def exact_example(word: object) -> str | None:
    for form in word_forms(word):
        if form in EXACT_EXAMPLES:
            return EXACT_EXAMPLES[form]
    return None


def source_example(word: object, source: dict[str, str]) -> str | None:
    for form in word_forms(word):
        if form in source:
            return source[form]
    return None


def generated_example(word: object, meaning: object) -> str:
    surface = surface_word(word)
    gloss = first_gloss(meaning)
    adj_cn = clean_adj_cn(gloss)

    if has_any(gloss, ("可口", "好吃", "美味")):
        return f"The soup tasted {surface}. (那碗湯喝起來很{adj_cn}。)"
    if has_any(gloss, ("噁心", "討厭")):
        return f"The old food smelled {surface}. (那份舊食物聞起來很{adj_cn}。)"
    if has_any(gloss, ("可信", "可能", "普通", "清楚", "貴重", "重要", "危險", "困難", "容易")) or gloss.endswith("的"):
        return f"The idea sounds {surface}. (這個想法聽起來很{adj_cn}。)"
    if has_any(gloss, ("灰色", "紅色", "藍色", "綠色", "白色", "黑色", "黃色", "紫色", "銀色", "顏色")):
        return f"The sky turned {surface} before sunset. (日落前天空變成{gloss}。)"
    if has_any(gloss, ("元", "分", "貨幣", "硬幣")):
        return f"She found a {surface} on the sidewalk. (她在人行道上找到一個{gloss}。)"
    if has_any(gloss, ("打", "十二")):
        return f"We bought a dozen eggs at the market. (我們在市場買了一打蛋。)"
    if has_any(gloss, ("餐", "菜", "麵", "飯", "肉", "豆", "蘿蔔", "麥片", "可可", "糖", "蛋糕", "食物", "水果", "果醬", "鮮奶油", "作物", "湯", "鹽", "紅酒", "番薯", "餃子", "葡萄", "番石榴", "漢堡", "蜂蜜", "檸檬", "芒果", "瓜", "堅果", "洋蔥", "木瓜", "桃子", "花生", "梨子", "胡椒", "鳳梨", "南瓜", "沙拉", "三明治", "醬汁", "點心", "萵苣", "生菜", "甜點", "甜甜圈", "橘子", "吐司")):
        return f"I had some {surface} for breakfast. (我早餐吃了一些{gloss}。)"
    if has_any(gloss, ("醫生", "老師", "學生", "司機", "廚師", "舞蹈家", "清潔工", "教練", "水手", "丈夫", "情侶", "夫妻", "顧客", "護士", "女王", "皇后", "先生", "太太", "女士", "牙醫", "鬥士", "漁夫", "紳士", "男性", "夥伴", "姪子", "外甥", "姪女", "外甥女", "乘客", "王子", "公主", "囚犯", "秘書", "青少年", "小偷", "竊賊", "訪客", "服務生", "青年", "年輕人", "兒", "女兒", "祖父", "祖母", "孫", "人", "者", "員", "師", "家", "工")):
        return f"The {surface} helped the students after class. (那位{gloss}下課後幫助學生。)"
    if has_any(gloss, ("烏鴉", "幼獸", "老鷹", "青蛙", "螃蟹", "鶴", "鵝", "蟑螂", "動物", "鳥", "魚", "狗", "貓", "蟲", "老鼠", "大鼠", "蛇", "龍", "蜻蜓", "狐狸", "豹", "蚊子", "蛾", "騾", "鸚鵡", "蝦", "蜘蛛", "驢", "恐龍", "怪物", "妖怪", "松鼠", "烏龜")):
        return f"We saw {article_for(surface)} {surface} near the river. (我們在河邊看到一隻{gloss}。)"
    if has_any(gloss, ("雲", "霧", "氣候", "空氣", "風", "雨", "雪", "海岸", "森林", "河", "天空")):
        return f"The {surface} changed before evening. (傍晚前{gloss}有了變化。)"
    if has_any(gloss, ("店", "館", "校", "公司", "餐廳", "公寓", "俱樂部", "鄉間", "郡", "海岸", "角落", "山丘", "房子", "法庭", "球場", "村", "城市", "地方", "池塘", "游泳池", "水池", "鐵路", "城鎮", "車庫", "高速公路", "島嶼", "王國", "小路", "港口", "監獄", "洗手間", "藥房", "入口", "行星", "地鐵", "寺廟", "廁所", "塔", "軌道", "隧道", "假期", "山谷")):
        return f"We visited the {surface} after school. (我們放學後去了那個{gloss}。)"
    if has_any(gloss, ("假日", "小時", "黎明", "一天", "時間")):
        return f"We rested during the {surface}. (我們在那段{gloss}期間休息。)"
    if has_any(gloss, ("課程", "科目", "考試", "危險", "死亡", "夢想", "感情", "心情", "勇氣", "衝突", "祝賀", "藝術", "西洋棋", "歷史", "主意", "功課", "興趣", "犯罪", "危機", "文化", "習俗", "損壞", "資料", "數據", "辯論", "債", "決定", "音樂", "詩歌", "詩集", "排", "爭吵", "拍賣", "銷售", "恐懼", "自己", "句子", "形狀", "外形", "景象", "靈魂", "演講", "運動", "故事", "思想", "音調", "總額", "總數", "旅遊", "麻煩", "婚禮", "戲劇", "描繪", "素描", "責任", "節慶", "發燒", "驚嚇", "地理", "高爾夫", "成長", "嗜好", "幽默", "飢餓", "收入", "獨立", "企業", "行業", "例子", "實例", "邀請", "請帖", "項目", "領導", "長度", "遺失", "運氣", "幸運", "禮貌", "方式", "婚姻", "大量", "意義", "意思", "手段", "會議", "聚會", "記憶", "回憶", "訊息", "方法", "百萬", "減去", "負號", "缺點", "動作", "小說", "樂趣", "愉悅", "詩", "毒藥", "姿勢", "出席", "得意", "原則", "進步", "發展", "測驗", "小考", "現實", "實際情況", "關係", "親戚", "財富", "安全", "秘密", "選拔", "學期", "安頓", "定居", "衝撞", "安靜", "解決方案", "來源", "速度", "拼字", "精神", "流感", "爵士樂", "華語", "片語", "一堆", "一疊", "污點", "供應", "鞦韆", "符號", "天賦", "目標", "青少年時期", "標題", "職稱", "主題", "貿易", "傳統", "交通", "考驗", "真相", "字彙", "勝利", "耳語", "擴大", "寬度", "創傷")):
        return f"We talked about {surface} in class. (我們在課堂上談到{gloss}。)"
    if has_any(gloss, ("確認", "包含", "咳嗽", "申請", "參加", "逮捕", "做", "複製")):
        return f"Please {surface} it carefully before class. (請在上課前仔細{gloss}它。)"
    if has_any(gloss, ("嘴唇", "尾巴", "肚子", "眉毛", "臀部", "關節", "指甲", "手心", "器官", "胃", "大拇指", "腰部", "翅膀", "傷口")):
        return f"He hurt his {surface} during practice. (他在練習時弄傷了{gloss}。)"
    if has_any(gloss, ("電腦", "日曆", "蠟燭", "粉筆", "梳子", "叉子", "籠子", "硬幣", "筷子", "布", "煤", "棉", "黏土", "手推車", "烹調器具", "捆", "束", "盤", "碟", "眼鏡", "洋娃娃", "草", "草地", "洞", "英寸", "蠟筆", "泥巴", "報紙", "橡膠", "橡皮", "鋸子", "種子", "翹翹板", "肥皂", "沙發", "湯匙", "階梯", "票", "領帶", "輪胎", "玩具", "T 恤", "屋頂", "根", "繩子", "玫瑰", "鼓", "抽屜", "洋裝", "藥品", "烘乾機", "信封", "橡皮擦", "手電筒", "冷凍庫", "吉他", "鐵鎚", "手帕", "暖氣", "樂器", "夾克", "外套", "吉普車", "燈籠", "蓋子", "圓木", "雜誌", "口罩", "面具", "地墊", "材料", "物質", "藥", "旋律", "金屬", "鏡子", "摩托車", "捷運", "項鍊", "筆記本", "包裹", "繪畫", "平底鍋", "管子", "煙斗", "明信片", "壺", "印表機", "獎品", "獎金", "錢包", "謎題", "難題", "長方形", "直尺", "樣品", "剪刀", "螢幕", "架子", "殼", "絲綢", "溜滑梯", "閃電", "長笛", "液體", "蒸汽", "鋼", "棒狀物", "爐子", "吸管", "細繩", "西裝", "坦克車", "教科書", "塔", "陷阱", "寶藏", "三角形", "小號", "管", "制服", "武器", "輪子", "電線", "木偶", "正方形")):
        return f"The {surface} is on the table. (那個{gloss}在桌上。)"
    if has_any(gloss, ("再見", "謝謝", "恭喜")):
        return f"She smiled and said {surface}. (她微笑著說{gloss}。)"

    return f"The {surface} appeared in the story. (故事中出現了{gloss}。)"


def choose_example(word: object, meaning: object, old_example: object, current_example: object, source: dict[str, str]) -> str:
    exact = exact_example(word)
    if exact:
        return exact

    current = str(current_example or "").strip()
    if current and not is_bad(current) and not is_generic(current):
        return current

    old = str(old_example or "").strip()
    if old and not is_bad(old) and not is_generic(old):
        return old

    found = source_example(word, source)
    if found and not is_bad(found) and not is_generic(found):
        return found

    return generated_example(word, meaning)


def main() -> None:
    if not BACKUP_PATH.exists():
        shutil.copy2(WORKBOOK_PATH, BACKUP_PATH)

    source = load_source_examples()
    wb = load_workbook(WORKBOOK_PATH)
    for sheet_name in ("Level 1", "Level 2"):
        ws = wb[sheet_name]
        ws.cell(1, 5).value = "新例句"
        for row in range(2, ws.max_row + 1):
            ws.cell(row, 5).value = choose_example(
                ws.cell(row, 2).value,
                ws.cell(row, 3).value,
                ws.cell(row, 4).value,
                ws.cell(row, 5).value,
                source,
            )

    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
