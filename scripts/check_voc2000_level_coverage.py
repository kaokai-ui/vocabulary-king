from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent


def normalize(text: str) -> str:
    text = text.strip().lower().replace("／", "/").replace("’", "'")
    return re.sub(r"\s+", " ", text)


def candidates(entry: str) -> list[str]:
    entry = normalize(entry)
    base = re.sub(r"\s*\([^)]*\)", "", entry).strip()
    options = [entry, base]
    options.extend(part.strip() for part in base.split("/") if part.strip())
    return list(dict.fromkeys(option for option in options if option))


voc_rows = json.loads((REPO_ROOT / "outputs" / "voc2000_meanings.json").read_text(encoding="utf-8"))
voc_map = {normalize(row["word"]): row["meaning"] for row in voc_rows}

wb = load_workbook(
    REPO_ROOT / "outputs" / "ceec_vocab_levels" / "詞彙分級表_levels_附意思例句.xlsx",
    data_only=True,
)
for sheet_name in ["Level 1", "Level 2"]:
    ws = wb[sheet_name]
    missing = 0
    found = 0
    samples = []
    for row in range(2, ws.max_row + 1):
        if ws[f"C{row}"].value:
            continue
        missing += 1
        word = str(ws[f"B{row}"].value).strip()
        for candidate in candidates(word):
            if candidate in voc_map:
                found += 1
                if len(samples) < 12:
                    samples.append((word, candidate, voc_map[candidate]))
                break
    print(sheet_name, "missing", missing, "found_in_voc2000", found)
    for sample in samples:
        print(" ", sample)
